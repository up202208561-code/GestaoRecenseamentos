import pandas as pd
import openpyxl
import tempfile

from copy import copy
from formulas import FORMULAS
from dados import CAMPOS
from io import BytesIO

def abrir_excel(excel_bytes):
    """
    Abre um Excel que está em memória (bytes).
    """

    return openpyxl.load_workbook(
        BytesIO(excel_bytes),
        keep_vba=True
    )


def ler_recenseamentos(ficheiro):

    """
    Lê a folha Recenseamentos e devolve um DataFrame
    """

    if isinstance(ficheiro, bytes):
        ficheiro = BytesIO(ficheiro)

    dados = pd.read_excel(
        ficheiro,
        sheet_name="Recenseamentos",
        header=None
    )

    dados = dados.iloc[5:]

    projetos = pd.DataFrame()

    for campo in CAMPOS:
        projetos[campo["campo"]] = dados.iloc[:, campo["coluna"] - 1]

    projetos = projetos.dropna(subset=["RefObra"])

    return projetos


def pesquisar_projetos(projetos, texto):
    """
    Pesquisa um texto em todas as colunas.
    """

    texto = texto.lower()

    return projetos[
        projetos.astype(str)
        .apply(
            lambda coluna:
                coluna.str.lower().str.contains(texto, na=False)
        )
        .any(axis=1)
    ]


def guardar_estado(ficheiro, ref_obra, novo_estado):
    """
    Atualiza o estado de uma obra.
    """

    temp = tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".xlsm"
    )

    temp.write(ficheiro.getvalue())
    temp.close()

    wb = openpyxl.load_workbook(
        temp.name,
        keep_vba=True
    )

    ws = wb["Recenseamentos"]

    for linha in range(6, ws.max_row + 1):

        referencia = ws.cell(
            row=linha,
            column=3
        ).value

        if referencia == ref_obra:

            ws.cell(
                row=linha,
                column=4
            ).value = novo_estado

            break

    wb.save(temp.name)

    return temp.name
    


def procurar_linha_projeto(ws, ref_obra):
    """
    Procura a linha correspondente a uma referência de obra.
    """

    for linha in range(6, ws.max_row + 1):

        if ws.cell(row=linha, column=3).value == ref_obra:
            return linha

    return None


def ler_projeto(ws, linha):
    """
    Lê todos os campos definidos em CAMPOS para uma determinada linha
    da folha Recenseamentos.

    Devolve um dicionário.
    """

    projeto = {}

    for campo in CAMPOS:

        valor = ws.cell(
            row=linha,
            column=campo["coluna"]
        ).value

        if campo["campo"] == "TaxaPenetracao" and valor is not None:
            valor *= 100

        projeto[campo["campo"]] = valor

    return projeto

def guardar_projeto(excel_bytes, ref_obra, dados_projeto):
    """
    Atualiza todos os campos editáveis de uma obra.
    dados_projeto é um dicionário:
    {
        "Concelho": "...",
        "Estado": "...",
        "Rua": "...",
        ...
    }
    """

    temp = tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".xlsm"
    )

    temp.write(
        excel_bytes
    )

    temp.close()

    wb = openpyxl.load_workbook(
        temp.name,
        keep_vba=True
    )

    ws = wb["Recenseamentos"]

    linha = procurar_linha_projeto(
        ws,
        ref_obra
    )

    if linha is None:
        raise Exception("Projeto não encontrado.")

    for campo in CAMPOS:

        if not campo["editavel"]:
            continue

        nome = campo["campo"]

        if nome in dados_projeto:
                
            valor = dados_projeto[nome]
        
            if nome == "TaxaPenetracao":
                valor = valor / 100
            
            ws.cell(
                row=linha,
                column=campo["coluna"]
            ).value = valor

        for coluna, formula in FORMULAS.items():
            ws.cell(
                row=linha,
                column=coluna
            ).value = formula.format(r=linha)

    wb.save(temp.name)

    with open(temp.name, "rb") as f:
        return f.read()

def criar_projeto(excel_bytes, dados_projeto):
    """
    Cria uma nova obra mantendo fórmulas, estilos e formatação.
    Se não existir nenhuma obra, utiliza a linha 6 como modelo.
    """

    temp = tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".xlsm"
    )

    temp.write(excel_bytes)
    temp.close()

    wb = openpyxl.load_workbook(
        temp.name,
        keep_vba=True
    )

    ws = wb["Recenseamentos"]

    # ---------------------------------
    # Encontrar última obra existente
    # ---------------------------------

    ultima_linha = None

    for linha in range(6, ws.max_row + 1):

        if ws.cell(row=linha, column=3).value not in (None, ""):
            ultima_linha = linha

    # ---------------------------------
    # Definir linha modelo e nova linha
    # ---------------------------------

    if ultima_linha is None:

        # Primeira obra
        nova_linha = 6

        # Se existir uma linha 6 no template, usa-a como modelo
        if ws.max_row >= 6:
            linha_modelo = 6
        else:
            linha_modelo = None

    else:

        linha_modelo = ultima_linha
        nova_linha = ultima_linha + 1

    # ---------------------------------
    # Copiar formatação
    # ---------------------------------

    if linha_modelo is not None:

        ws.row_dimensions[nova_linha].height = \
            ws.row_dimensions[linha_modelo].height

        for col in range(1, ws.max_column + 1):

            origem = ws.cell(
                row=linha_modelo,
                column=col
            )

            destino = ws.cell(
                row=nova_linha,
                column=col
            )

            destino.font = copy(origem.font)
            destino.fill = copy(origem.fill)
            destino.border = copy(origem.border)
            destino.alignment = copy(origem.alignment)
            destino.protection = copy(origem.protection)
            destino.number_format = copy(origem.number_format)

    # ---------------------------------
    # Limpar toda a linha
    # ---------------------------------

    for col in range(1, ws.max_column + 1):

        ws.cell(
            row=nova_linha,
            column=col
        ).value = None

    # ---------------------------------
    # Escrever campos manuais
    # ---------------------------------

    for campo in CAMPOS:

        if not campo["editavel"]:
            continue

        nome = campo["campo"]

        if nome not in dados_projeto:
            continue

        valor = dados_projeto[nome]

        if nome == "TaxaPenetracao":
            valor = valor / 100

        ws.cell(
            row=nova_linha,
            column=campo["coluna"]
        ).value = valor

    # ---------------------------------
    # Escrever fórmulas
    # ---------------------------------

    for coluna, formula in FORMULAS.items():

        ws.cell(
            row=nova_linha,
            column=coluna
        ).value = formula.format(r=nova_linha)

    wb.save(temp.name)

    with open(temp.name, "rb") as f:
        return f.read()

